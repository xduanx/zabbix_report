import copy
import json
import re
import ssl
import time
import urllib
from urllib import request
import os

import chardet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#proper ip replaced
zabbix_url = "http://192.168.0.1/zabbix/api_jsonrpc.php"
headers = {
    "Content-Type": "application/json-rpc"
}


def get_token():
    if os.path.exists(BASE_DIR + "/zabbix_token.txt"):
        with open(BASE_DIR + "/zabbix_token.txt", "r") as f:
            zabbix_token = f.read()
    else:
        post_data = {
            "jsonrpc": "2.0",
            "method": "user.login",
            "params": {
                "user": "Admin",
                # proper password replaced
                "password": "Admin"
            },
            "id": 1
        }
        rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                              method='POST')
        response = request.urlopen(rqs)
        dict_access_token = json.loads(response.read())
        zabbix_token = dict_access_token["result"]
        with open(BASE_DIR + "/zabbix_token.txt", "w") as f1:
            f1.write(zabbix_token)
    return zabbix_token


def get_hosts():
    post_data = {
        "jsonrpc": "2.0",
        "method": "host.get",
        "params": {
            "output": [
                "hostid",
                "host"
            ],
            "selectGroups": [
                "groupid",
                "name"
            ],
        },
        "id": 2,
        "auth": get_token()
    }
    rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                          method='POST')
    response = request.urlopen(rqs)
    return json.loads(response.read())["result"]


def get_groups():
    post_data = {
        "jsonrpc": "2.0",
        "method": "hostgroup.get",
        "params": {
            "output": "extend",
            "real_hosts": "",
        },
        "auth": get_token(),
        "id": 1
    }
    rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                          method='POST')
    response = request.urlopen(rqs)
    return json.loads(response.read())["result"]


def get_host_info(groupid):
    post_data = {
        "jsonrpc": "2.0",
        "method": "host.get",
        "params": {
            "output": [
                "hostid",
                "host"
            ],
            "groupids": groupid,
        },
        "id": 2,
        "auth": get_token()
    }
    rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                          method='POST')
    response = request.urlopen(rqs)
    return json.loads(response.read())["result"]


def get_item_value(hostid_query, item_query):
    post_data = {
        "jsonrpc": "2.0",
        "method": "item.get",
        "params": {
            "output": ["itemid", "hostid", "key_", "lastvalue", "lastclock", "units"],
            "hostids": hostid_query,
            "search": {
                "key_": item_query
            },
            "sortfield": "key_"
        },
        "auth": get_token(),
        "id": 1
    }
    rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                          method='POST')
    try_times = 1
    while try_times < 4:
        try:
            response = request.urlopen(rqs)
            return json.loads(response.read())["result"]
        except Exception as e:
            print(e)
            time.sleep(5)
            print(str(try_times) + " times reconnecting ...")
            try_times += 1


def get_hosts_groups_info(groups):
    hosts_groups_info = []
    for i in groups:
        for j in get_groups():
            if j["name"] == i:
                host_group_info = get_host_info(j["groupid"])
                for k in host_group_info:
                    k.update({"group": i, "groupid": j["groupid"]})
                hosts_groups_info += host_group_info
                break
    return hosts_groups_info


def get_tablespace_info(groups):
    tablespace_info = []
    item_info = {"hostid": "", "host": "", "item": "", "total": "", "free": "", "used": "", "group": ""}

    groups = [h for h in groups if h.find("db") > 0]
    for i in get_hosts_groups_info(groups):
        tb_list = []
        host_tb_info = get_item_value(i["hostid"], "tb_size")
        for j in host_tb_info:
            tb_list.append(j["key_"].split(",")[2].rstrip("]"))
        tb_list = list(set(tb_list))
        for k in tb_list:
            item_info["hostid"] = i["hostid"]
            item_info["host"] = i["host"]
            item_info["item"] = k
            item_info["used"] = get_item_value(i["hostid"], "zabora[{$ORACLE_SID},tb_size," + k + "]")[0]["lastvalue"]
            item_info["free"] = get_item_value(i["hostid"], "zabora[{$ORACLE_SID},tb_free_size," + k + "]")[0][
                "lastvalue"]
            item_info["total"] = str(round(float(item_info["free"]) * 100 / (100 - float(item_info["used"])), 2))
            item_info["group"] = i["group"]
            print(item_info)
            tablespace_info.append(copy.deepcopy(item_info))
    return tablespace_info


def get_partition_info(groups):
    partition_info = []
    item_info = {"hostid": "", "host": "", "item": "", "total": "", "free": "", "used": "", "group": ""}

    for i in get_hosts_groups_info(groups):
        partition_list = []
        host_partition_info = get_item_value(i["hostid"], "vfs.fs.size")
        for j in host_partition_info:
            partition_list.append(j["key_"].split(",")[0].split("[")[1])
        partition_list = list(set(partition_list))
        partition_exclude = ["boot", "tmp", "opt", "usr", "var", "overlay", "docker", "containers"]

        for j2 in partition_exclude:
            '''
            partition_list = [j3 for j3 in partition_list if j3.find(j2) < 0]
            '''
            # another implementation for partition_list
            for j3 in partition_list:
                if j3.find(j2) >= 0:
                    partition_list.remove(j3)

        print(partition_list)
        for k in partition_list:
            item_info["hostid"] = i["hostid"]
            item_info["host"] = i["host"]
            item_info["item"] = k
            item_info["total"] = str(
                round(int(get_item_value(i["hostid"], "vfs.fs.size[" + k + ",total]")[0]["lastvalue"]) / 1073741824, 2))
            item_info["used"] = get_item_value(i["hostid"], "vfs.fs.size[" + k + ",pused]")[0]["lastvalue"]
            item_info["free"] = str(
                round((int(get_item_value(i["hostid"], "vfs.fs.size[" + k + ",total]")[0]["lastvalue"]) -
                       int(get_item_value(i["hostid"], "vfs.fs.size[" + k + ",used]")[0]["lastvalue"])) /
                      1073741824, 2))
            item_info["group"] = i["group"]
            print(item_info)
            partition_info.append(copy.deepcopy(item_info))
    return partition_info


def get_diskgroup_info(groups):
    diskgroup_info = []
    item_info = {"hostid": "", "host": "", "item": "", "total": "", "free": "", "used": "", "group": ""}

    groups = [h for h in groups if h.find("db") > 0]
    for i in get_hosts_groups_info(groups):
        diskgroup_list = []
        host_diskgroup_info = get_item_value(i["hostid"], "diskgroup[")
        if not host_diskgroup_info:
            continue
        for j in host_diskgroup_info:
            diskgroup_list.append(j["key_"].split(",")[2].rstrip("]"))
        diskgroup_list = list(set(diskgroup_list))

        for k in diskgroup_list:
            item_info["hostid"] = i["hostid"]
            item_info["host"] = i["host"]
            item_info["item"] = k
            item_info["used"] = \
                get_item_value(i["hostid"], "diskgroup[{$ORACLE_SID},diskgroup_used_size," + k + "]")[0]["lastvalue"]
            item_info["free"] = \
                get_item_value(i["hostid"], "diskgroup[{$ORACLE_SID},diskgroup_free_size," + k + "]")[0]["lastvalue"]
            item_info["total"] = str(round(float(item_info["free"]) * 100 / (100 - float(item_info["used"])), 2))
            item_info["group"] = i["group"]
            print(item_info)
            diskgroup_info.append(copy.deepcopy(item_info))
    return diskgroup_info


def get_cpu_info(groups):
    cpu_info = []
    item_info = {"hostid": "", "host": "", "item": "", "total": "", "free": "", "used": "", "group": ""}
    for i in get_hosts_groups_info(groups):
        item_info["hostid"] = i["hostid"]
        item_info["host"] = i["host"]
        item_info["item"] = "cpu"
        item_info["total"] = get_item_value(i["hostid"], "system.cpu.num")[0]["lastvalue"]
        item_info["free"] = ""
        item_info["used"] = get_item_value(i["hostid"], "system.cpu.load[all,avg15]")[0]["lastvalue"]
        item_info["group"] = i["group"]
        print(item_info)
        cpu_info.append(copy.deepcopy(item_info))
    return cpu_info


def write_excel(output, content_json):
    cell_color = PatternFill("solid", fgColor="E97659")
    title_color = PatternFill("solid", fgColor="ADADAD")
    alignment_center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))
    file_name = output
    wb = Workbook()
    for key, value in content_json.items():
        ws = wb.create_sheet(title=key, index=0)
        ws['A1'] = 'item'
        ws["B1"] = "total_space(GB)"
        ws["C1"] = "free_space(GB)"
        ws["D1"] = "used_percengate(%)"
        ws["E1"] = "host"
        ws["F1"] = "group"
        for i in range(2, len(value) + 1):
            ws["A" + str(i)] = value[i - 2]["item"]
            ws["B" + str(i)] = float(value[i - 2]["total"])
            if value[i - 2]["free"] == "":
                ws["C" + str(i)] = value[i - 2]["free"]
            else:
                ws["C" + str(i)] = float(value[i - 2]["free"])
            ws["D" + str(i)] = float(value[i - 2]["used"])
            ws["E" + str(i)] = value[i - 2]["host"]
            ws["F" + str(i)] = value[i - 2]["group"]

        for j in ['A', 'B', 'C', 'D', 'E', 'F']:
            for i in range(1, len(value) + 1):
                ws[j + str(i)].font = Font(size=9)
                ws[j + str(i)].alignment = alignment_center
                ws[j + str(i)].border = border_thin
            ws[j + "1"].fill = title_color
            ws.column_dimensions[j].width = 20.0
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = 'A1:F1'
        ws.auto_filter.add_filter_column(0, [])
    wb.save(file_name)


def get_web_scenario(hostid="14782"):
    post_data = {
        "jsonrpc": "2.0",
        "method": "httptest.get",
        "params": {
            "hostids": hostid,
            "output": "extend",
            "selectSteps": "extend",
        },
        "auth": get_token(),
        "id": 1
    }
    rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                          method='POST')
    response = request.urlopen(rqs)
    return json.loads(response.read())["result"]


def get_trigger(hostid):
    post_data = {
        "jsonrpc": "2.0",
        "method": "trigger.get",
        "params": {
            "hostids": hostid,
            "output": "extend",
            "selectFunctions": "extend"
        },
        "auth": get_token(),
        "id": 1
    }
    rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                          method='POST')
    response = request.urlopen(rqs)
    return json.loads(response.read())["result"]


class Url(object):
    def __init__(self, url_entry):
        self.url_entry = url_entry

    def get_group_url_ip(self):
        headers_pachong = {
            "User-Agent": "User-Agent: Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:69.0) Gecko/20100101 Firefox/69.0",
            "Content-Type": "application/x-www-form-urlencoded; charset=utf-8"
        }
        url_group_list = dict()
        for x in self.url_entry:
            ssl._create_default_https_context = ssl._create_unverified_context
            rq = urllib.request.Request(url=x, headers=headers_pachong)
            try:
                response = urllib.request.urlopen(rq)
            except Exception as e:
                print(x + " is 404")
                continue
            html = response.read()
            encode_type = chardet.detect(html)
            html = html.decode(encode_type['encoding'])
            reg = r'<td>[0-9]{1,}.[0-9]{1,}.[0-9]{1,}.[0-9]{1,}:[0-9]{1,}</td>'
            url_reg = re.compile(reg)
            url_list = re.findall(url_reg, html)
            url_ip_port = []
            for y in url_list:
                y = y.rstrip("</td>").lstrip("<td>")
                y = "http://" + y + "/"
                url_ip_port.append(y)
            # x = x.lstrip("http://").rstrip("/status").lstrip("https://")
            url_group_list.update({x: url_ip_port})
        return url_group_list


def nginx_sort():
    with open(BASE_DIR + "/nginx_new_dr.txt", "r") as url_list:
        url_group = []
        for k in url_list:
            url_group.append(k.rstrip("\n"))
    url_all_list = Url(url_group)
    entry_url_dic = url_all_list.get_group_url_ip()
    for k1, v1 in entry_url_dic.items():
        entry_url_dic[k1] = str(v1)
    list_value = set(entry_url_dic.values())
    dict_key_value = {}
    for u in list_value:
        dict_key_value.update({u: []})
    for k, v in entry_url_dic.items():
        if v in dict_key_value:
            dict_key_value[v].append(k)
    return dict_key_value


def nginx_sort_export_excel(output, content_json):
    cell_color = PatternFill("solid", fgColor="CCCCCC")
    title_color = PatternFill("solid", fgColor="ADADAD")
    alignment_center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))
    file_name = output
    wb = Workbook()
    ws = wb.create_sheet(title="nginx_list", index=0)
    ws['A1'] = 'num'
    ws['B1'] = 'nginx'
    ws['A1'].fill = title_color
    ws['B1'].fill = title_color
    ws['A1'].font = Font(size=9)
    ws['B1'].font = Font(size=9)
    y = 1
    z = 2
    for key, value in content_json.items():
        for x in value:
            ws['A' + str(z)] = str(y)
            ws['B' + str(z)] = x
            ws['A' + str(z)].font = Font(size=9)
            ws['B' + str(z)].font = Font(size=9)
            if y % 2:
                ws['A' + str(z)].fill = cell_color
                ws['B' + str(z)].fill = cell_color
            z = z + 1
        y = y + 1
    ws.column_dimensions['B'].width = 35.0
    wb.save(file_name)


def get_problems_backend_ip(ip):
    headers_pachong = {
        "User-Agent": "User-Agent: Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:69.0) Gecko/20100101 Firefox/69.0",
        "Content-Type": "application/x-www-form-urlencoded; charset=utf-8"
    }
    ssl._create_default_https_context = ssl._create_unverified_context
    rq = urllib.request.Request(url=ip, headers=headers_pachong)
    response = urllib.request.urlopen(rq)
    html = response.read()
    encode_type = chardet.detect(html)
    html = html.decode(encode_type['encoding'])
    reg = r'<td>[0-9]{1,}.[0-9]{1,}.[0-9]{1,}.[0-9]{1,}:[0-9]{1,}</td>\s+<td>down</td>\s+<td>.*</td>\s+<td>.*</td>'
    url_reg = re.compile(reg)
    url_list = re.findall(url_reg, html)
    url_list = [
        {"ip": i.split()[0].rstrip("</td>").lstrip("<td>"), "fall_counts": i.split()[3].rstrip("</td>").lstrip("<td>")}
        for i in url_list]
    return url_list


'''
def get_items_top(item_name, limit_num, group_id="5"):
    post_data = {
        "jsonrpc": "2.0",
        "method": "item.get",
        "params": {
            "output": "extend",
            "groupids": group_id,
            "search": {
                "key_": item_name
            },
            "sortfield": "name"
        },
        "auth": get_token(),
        "id": 1
    }
    rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                          method='POST')
    response = request.urlopen(rqs)

    top_limit_items = [{"host_id": "", "host_ip": "", "value": "0", "host_group": ""} for i in range(limit_num)]
    for x in json.loads(response.read())["result"]:
        if float(x["lastvalue"]) <= float(top_limit_items[limit_num - 1]["value"]):
            continue
        elif float(x["lastvalue"]) >= float(top_limit_items[0]["value"]):
            del top_limit_items[limit_num - 1]
            top_limit_items.insert(0,
                                   {"host_id": x["hostid"], "host_ip": "", "value": x["lastvalue"], "host_group": ""})
        else:
            for y in range(limit_num - 1):
                if float(top_limit_items[y]["value"]) >= float(x["lastvalue"]) >= float(
                        top_limit_items[y + 1]["value"]):
                    del top_limit_items[limit_num - 1]
                    top_limit_items.insert(y + 1, {"host_id": x["hostid"], "host_ip": "", "value": x["lastvalue"],
                                                   "host_group": ""})
                    break
    return top_limit_items
'''

if __name__ == '__main__':
    start_time = int(time.time())
    pid = os.getpid()

    process_info = {"start_time": start_time, "pid": pid}
    with open(BASE_DIR + "/query_data_api_process_info.txt", "w") as query_data_api_process_info:
        query_data_api_process_info.write(json.dumps(process_info, indent=4, ensure_ascii=False))

    groups_selected = ['oracle_production_os', 'oracle_production_db', 'os_production', 'os_test', 'oracle_notused_db',
                       'oracle_notused_os', 'oracle_test_db', 'oracle_test_os']

    tablespace = get_tablespace_info(groups_selected)
    with open(BASE_DIR + "/tablespace.json", "w") as f:
        f.write(json.dumps(tablespace, indent=4, ensure_ascii=False))

    diskgroup = get_diskgroup_info(groups_selected)
    with open(BASE_DIR + "/diskgroup.json", "w") as f:
        f.write(json.dumps(diskgroup, indent=4, ensure_ascii=False))

    partition = get_partition_info(groups_selected)
    with open(BASE_DIR + "/partition.json", "w") as f:
        f.write(json.dumps(partition, indent=4, ensure_ascii=False))

    cpu = get_cpu_info(groups_selected)
    with open(BASE_DIR + "/cpu.json", "w") as f:
        f.write(json.dumps(cpu, indent=4, ensure_ascii=False))

    with open(BASE_DIR + "/report.json", "w") as f:
        f.write(json.dumps({"tablespace": tablespace, "diskgroup": diskgroup, "partition": partition, "cpu": cpu},
                           indent=4, ensure_ascii=False))

    with open(BASE_DIR + "/report.json", "r") as f:
        report_json = json.loads(f.read())
        write_excel(BASE_DIR + "/zabbix_report_" + str(time.strftime('%Y%m%d')) + ".xlsx", report_json)

    end_time = int(time.time())
    os.remove(BASE_DIR + "/query_data_api_process_info.txt")
    print("time:" + str((end_time - start_time) / 60) + "min")
