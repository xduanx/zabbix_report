import copy
import json
import re
import ssl
import time
import urllib
from urllib import request
import os

import chardet
import openpyxl
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from zabbix_items.models import Hostip, Hostinfomation

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# ip replaced
zabbix_url = "http://192.168.1.1/zabbix/api_jsonrpc.php"
headers = {
    "Content-Type": "application/json-rpc"
}


def get_token():
    if os.path.exists(BASE_DIR + "/zabbix_token.txt"):
        with open(BASE_DIR + "/zabbix_token.txt", "r") as f:
            zabbix_token = f.read()
    else:
        post_data = {
            "jsonrpc": "2.0",
            "method": "user.login",
            "params": {
                "user": "Admin",
                # password replaced
                "password": "Admin"
            },
            "id": 1
        }
        rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                              method='POST')
        response = request.urlopen(rqs)
        dict_access_token = json.loads(response.read())
        zabbix_token = dict_access_token["result"]
        with open(BASE_DIR + "/zabbix_token.txt", "w") as f1:
            f1.write(zabbix_token)
    return zabbix_token


def get_hosts():
    post_data = {
        "jsonrpc": "2.0",
        "method": "host.get",
        "params": {
            "output": "extend"
        },
        "id": 2,
        "auth": get_token()
    }
    rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                          method='POST')
    response = request.urlopen(rqs)
    return json.loads(response.read())["result"]


def zichan_guanli(request):
    show_host = Hostip.objects.all()
    return render(request, "cmdb.html", {"show_host": show_host})


'''
for i in get_hosts():
    newip = Hostip(ip=i['host'], hostname="")
    newip.save()
    print(i['host'])
'''


def add_hostinfomation():
    hosts = Hostip.objects.all()
    host_info = get_hosts()
    for i in host_info:
        print(i)

    '''
    for i in hosts:
        print(i.hostid)
        newhostinfo = Hostinfomation(hostid=i.hostid, sysadm_password="Admin", 
                                     zjgl_password="ZwWw93#CJBnBBX6", cwgl_password="Chf#Pq9Y2kgCpm6",
                                     root_password="RgnJMNqkY2PB#Ht")
        newhostinfo.save()
    '''


add_hostinfomation()


def get_templates(hostid):
    post_data = {
        "jsonrpc": "2.0",
        "method": "template.get",
        "params": {
            "hostids": hostid,
            "output": "extend"
        },
        "auth": get_token(),
        "id": 1
    }
    rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                          method='POST')
    response = request.urlopen(rqs)
    return json.loads(response.read())["result"]


def export_excel(output_file):
    cell_color = PatternFill("solid", fgColor="CCCCCC")
    title_color = PatternFill("solid", fgColor="ADADAD")
    alignment_center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))

    ip_range_list = []
    ip_list = []
    ip_range_dict = {}
    all_hosts_info = get_hosts()
    for i in all_hosts_info:
        ip = i["host"].split(".")
        if len(ip) < 4:
            continue
        ip_list.append(i["host"])
        del ip[3]
        ip_range_list.append(".".join(ip))
    ip_range_list = list(set(ip_range_list))

    for i in ip_range_list:
        ip_range_dict.update({i + ".": []})
    for k, v in ip_range_dict.items():
        for i in ip_list:
            if i.startswith(k):
                v.append(i)

    workbook = Workbook()
    sheet_ip_range_list = workbook.active
    sheet_ip_range_list.title = "ip_range_list"
    sheet_ip_range_list["A1"] = "ip_range"
    sheet_ip_range_list["B1"] = "主机数量"
    sheet_ip_range_list["C1"] = "备注"
    for i in ['A', 'B', 'C']:
        sheet_ip_range_list[i + "1"].font = Font(size=9)
        sheet_ip_range_list.column_dimensions[i].width = 15.0

    sheet_ip_range_list.freeze_panes = 'A2'
    sheet_ip_range_list.auto_filter.ref = 'A1:C1'
    sheet_ip_range_list.auto_filter.add_filter_column(0, [])

    row_num = 2
    for k, v in ip_range_dict.items():
        sheet_ip_range_list["A" + str(row_num)] = k + "0"
        sheet_ip_range_list["B" + str(row_num)] = len(v)

        sheet_hosts = workbook.create_sheet(k + "0")
        sheet_hosts["A1"] = "ip"
        sheet_hosts["B1"] = "oracle模板"
        sheet_hosts["C1"] = "os模板"
        sheet_hosts["D1"] = "ping模板"
        sheet_hosts["E1"] = "机器状态(enable/disable)"
        sheet_hosts["F1"] = "备注"
        for i in ['A', 'B', 'C', 'D', 'E', 'F']:
            sheet_hosts[i + "1"].font = Font(size=9)
            sheet_hosts.column_dimensions[i].width = 15.0
        sheet_hosts_row_num = 2
        for i in v:
            sheet_hosts["A" + str(sheet_hosts_row_num)] = i
            sheet_hosts["A" + str(sheet_hosts_row_num)].font = Font(size=9)
            sheet_hosts_row_num = sheet_hosts_row_num + 1

        sheet_ip_range_list["A" + str(row_num)].hyperlink = "#" + k + "0" + "!" + "A2"

        sheet_hosts.freeze_panes = 'A2'
        sheet_hosts.auto_filter.ref = 'A1:F1'
        sheet_hosts.auto_filter.add_filter_column(0, [])

        row_num = row_num + 1

    workbook.save(output_file)

    wb = openpyxl.load_workbook(output_file)
    for k, v in ip_range_dict.items():
        sheet = wb[k + "0"]
        row_count = len(list(sheet.rows))
        for i in range(2, row_count + 1):
            host_templates = []
            host_disable = 1
            for x in all_hosts_info:
                if x["host"] == sheet["A" + str(i)].value:
                    host_disable = int(x["disable_until"])
                    for y in get_templates(x["hostid"]):
                        host_templates.append(y["templateid"])
                    break
            print(host_templates)
            if '10359' in host_templates:
                sheet["B" + str(i)] = 1
                sheet["B" + str(i)].font = Font(size=9)
            sheet["C" + str(i)] = 1
            sheet["C" + str(i)].font = Font(size=9)
            sheet["D" + str(i)] = 1
            sheet["D" + str(i)].font = Font(size=9)
            if host_disable:
                sheet["E" + str(i)] = "disable"
                sheet["E" + str(i)].font = Font(size=9)
            else:
                sheet["E" + str(i)] = "enable"
                sheet["E" + str(i)].font = Font(size=9)

    wb.save(output_file)


def export_excel_49_range_2(output_file, input_file):
    wb = openpyxl.load_workbook(output_file)
    sheet = wb["Sheet1"]
    num = 2
    with open(input_file, "r") as log_file:
        for x in log_file:
            print(x.rstrip("\n").split(","))
            ip_status = x.rstrip("\n").split(",")
            sheet["A" + str(num)] = ip_status[0]
            if "no" in ip_status[1] and "no" in ip_status[2] and "no" in ip_status[3]:
                sheet["E" + str(num)] = "ping不通, 22不通, 3389不通"
            if "yes" in ip_status[1] and "no" in ip_status[2]:
                sheet["E" + str(num)] = "ping通, 22不通"
            if "yes" in ip_status[2] and "no" in ip_status[3] and "no" in ip_status[4] and "no" in ip_status[5]:
                sheet["E" + str(num)] = "zjgl, cwgl, root密码错误, 且不在已记录的资产表内"

            num = num + 1
        wb.save(output_file)


# export_excel_49_range_2(BASE_DIR + "/ip.xlsx", BASE_DIR + "/ip.result.log")
# export_excel(BASE_DIR + "/zabbix_host_report.xlsx")

'''
'name': 'Template OS Linux by Zabbix agent', 'flags': '0', 'templateid': '10001',
'name': 'Template Module ICMP Ping', 'flags': '0', 'templateid': '10186',
10359 | Template DB Oracle Agent V1.0
'''
