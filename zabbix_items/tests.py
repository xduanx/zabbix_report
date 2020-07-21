import json
import queue
import re
import ssl
import threading
import urllib
import random
from urllib import request

import chardet
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
import pymysql
import datetime, time
import csv
import os
from zabbix_items import query_data_api

# print(str(int(time.time())))
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
export_file = "./zabbix_report_" + str(time.strftime('%Y%m%d')) + ".xlsx"
cell_color = PatternFill("solid", fgColor="FF0000")
tb_all_value = []
diskgroup_all_value = []
diskspace_all_value = []


def csv_to_dict(path):
    with open(path, "r") as f1:
        return [{k: v for k, v in row.items()} for row in csv.DictReader(f1, skipinitialspace=True)]


def create_excel():
    file_name = export_file
    if os.path.exists(file_name):
        unixtime_value = str(int(time.time()))
        os.rename(file_name, file_name + "." + unixtime_value)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "tablespace"
    ws1['A1'] = 'tablespace_name'
    ws1["B1"] = "total_space(GB)"
    ws1["C1"] = "free_space(GB)"
    ws1["D1"] = "used_percengate(%)"
    ws1["E1"] = "IP"
    ws1["F1"] = "hostname"
    ws1["G1"] = "project_name"
    for j in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws1[j + "1"].font = Font(size=9)
        ws1.column_dimensions[j].width = 15.0
    ws2 = wb.create_sheet("diskgroup")
    ws2["A1"] = "diskgroup_name"
    ws2["B1"] = "total_space(GB)"
    ws2["C1"] = "free_space(GB)"
    ws2["D1"] = "used_percengate(%)"
    ws2["E1"] = "IP"
    ws2["F1"] = "hostname"
    ws2["G1"] = "project_name"
    for j in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws2[j + "1"].font = Font(size=9)
        ws2.column_dimensions[j].width = 15.0
    ws3 = wb.create_sheet("diskspace")
    ws3["A1"] = "partitation_name"
    ws3["B1"] = "total_space(GB)"
    ws3["C1"] = "free_space(GB)"
    ws3["D1"] = "used_percengate(%)"
    ws3["E1"] = "IP"
    ws3["F1"] = "hostname"
    ws3["G1"] = "project_name"
    for j in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws3[j + "1"].font = Font(size=9)
        ws3.column_dimensions[j].width = 15.0
    wb.save(file_name)


def item_query(ip, item_name, item_name_exclude="[[:blank:]]", table_name="history"):
    #ip and password replaced
    conn = pymysql.connect(host="192.168.1.1", user="zabbix_report", password="Admin", database="zabbix")
    cr = conn.cursor()
    query_date = datetime.date.today() + datetime.timedelta(-1)
    query_host = ip
    query_item = item_name
    query_exclude_item = item_name_exclude
    sql = "SELECT items.key_, interface.ip, hosts.host, a.value, a.clock " \
          "FROM " + table_name + " as a " \
                                 "INNER JOIN items " \
                                 "INNER JOIN hosts " \
                                 "INNER JOIN interface " \
                                 "INNER JOIN " + table_name + " as b " \
                                                              "ON a.itemid=items.itemid AND items.hostid=hosts.hostid AND items.hostid=interface.hostid AND a.itemid=b.itemid " \
                                                              "WHERE a.clock > UNIX_TIMESTAMP(\'" + str(
        query_date) + "\') AND items.key_ like \'%" + query_item + "%\' AND items.key_ not regexp \'" + query_exclude_item + "\' AND interface.ip=\'" + query_host + "\' " \
                                                                                                                                                                     "group by items.key_,interface.ip,hosts.host, a.value, a.clock having a.clock=max(b.clock) order by key_"
    cr.execute(sql)
    item_info = cr.fetchall()
    return item_info


def tablespace(ip, project_name):
    ip = ip
    tb_free_size = "tb_free_size"
    percentage_used = "tb_size"
    project_info = project_name

    tb_free_size_info = item_query(ip, tb_free_size)
    percentage_used_info = item_query(ip, percentage_used)
    tb_num = len(tb_free_size_info)

    for i in range(1, tb_num + 1):
        tb_name_value = tb_free_size_info[i - 1][0].split(",")[2].rstrip("]")
        tb_total_size_value = round(float(tb_free_size_info[i - 1][3]) * 100 / (100 - percentage_used_info[i - 1][3]),
                                    2)
        tb_free_size_value = tb_free_size_info[i - 1][3]
        percentage_used_value = percentage_used_info[i - 1][3]
        ip_value = ip
        hostname_value = tb_free_size_info[i - 1][2]
        project_value = project_info
        tb_value = [
            tb_name_value, tb_total_size_value, tb_free_size_value, percentage_used_value, ip_value, hostname_value,
            project_value]
        print(tb_value)
        tb_all_value.append(tb_value)


def diskspace(ip, project_name):
    used_size_item = "vfs.fs.size%used"
    total_size_item = "vfs.fs.size%total"
    percentage_used_item = "vfs.fs.size%pused"

    used_size_record = item_query(ip, used_size_item, "/boot|/home|/var|/tmp", "history_uint")
    total_size_record = item_query(ip, total_size_item, "/boot|/home|/var|/tmp", "history_uint")
    percentage_used_record = item_query(ip, percentage_used_item, "/boot|/home|/var|/tmp", "history")
    filesystem_num = len(used_size_record)

    for i in range(1, filesystem_num + 1):
        filesystem_name_value = used_size_record[i - 1][0].split(",")[0].lstrip("vfs.fs.size[")
        filesystem_total_size_value = round(int(total_size_record[i - 1][3]) / 1073741824, 2)
        filesystem_free_size_value = round(
            (int(total_size_record[i - 1][3]) - int(used_size_record[i - 1][3])) / 1073741824, 2)
        filesystem_percentage_used_value = percentage_used_record[i - 1][3]
        ip_value = ip
        hostname_value = used_size_record[0][2]
        project_value = project_name
        filesystem_list = [filesystem_name_value, filesystem_total_size_value, filesystem_free_size_value,
                           filesystem_percentage_used_value, ip_value, hostname_value, project_value]
        print(filesystem_list)
        diskspace_all_value.append(filesystem_list)


def diskgroup(ip, project_name):
    percentage_used_item = "diskgroup_used_size"
    free_size_item = "diskgroup_free_size"

    percentage_used_record = item_query(ip, percentage_used_item)
    free_size_record = item_query(ip, free_size_item)
    diskgroup_num = len(percentage_used_record)

    for i in range(1, diskgroup_num + 1):
        diskgroup_name_value = percentage_used_record[i - 1][0].split(",")[2].rstrip("]")
        diskgroup_free_size_value = round(float(free_size_record[i - 1][3]), 2)
        diskgroup_percentage_used_value = percentage_used_record[i - 1][3]
        diskgroup_total_size_value = round(
            diskgroup_free_size_value * 100 / (100 - float(diskgroup_percentage_used_value)), 2)
        ip_value = ip
        hostname_value = percentage_used_record[i - 1][2]
        project_value = project_name
        diskgroup_list = [diskgroup_name_value, diskgroup_total_size_value, diskgroup_free_size_value,
                          diskgroup_percentage_used_value, ip_value, hostname_value, project_value]
        print(diskgroup_list)
        diskgroup_all_value.append(diskgroup_list)


def tablespace_query(query_list):
    print("tablespace report, waiting ......")
    for i in query_list:
        tablespace(i["host_ip"], i["project_info"])
    row_num = 2
    workbook = openpyxl.load_workbook(export_file)
    sheet = workbook.get_sheet_by_name("tablespace")
    for i in tb_all_value:
        sheet["A" + str(row_num)] = i[0]
        sheet["A" + str(row_num)].font = Font(size=9)
        sheet["B" + str(row_num)] = i[1]
        sheet["B" + str(row_num)].font = Font(size=9)
        sheet["C" + str(row_num)] = i[2]
        sheet["C" + str(row_num)].font = Font(size=9)
        sheet["D" + str(row_num)] = i[3]
        sheet["D" + str(row_num)].font = Font(size=9)
        if float(sheet["D" + str(row_num)].value) > 80:
            for cell_list in sheet["A" + str(row_num): "G" + str(row_num)]:
                for cell in cell_list:
                    cell.fill = cell_color
        sheet["E" + str(row_num)] = i[4]
        sheet["E" + str(row_num)].font = Font(size=9)
        sheet["F" + str(row_num)] = i[5]
        sheet["F" + str(row_num)].font = Font(size=9)
        sheet["G" + str(row_num)] = i[6]
        sheet["G" + str(row_num)].font = Font(size=9)
        row_num = row_num + 1
    workbook.save(export_file)
    print("tablespace report, success")


def diskspace_query(query_list):
    print("diskspace report, waiting ......")
    for i in query_list:
        diskspace(i["host_ip"], i["project_info"])
    row_num = 2
    workbook = openpyxl.load_workbook(export_file)
    sheet = workbook.get_sheet_by_name("diskspace")
    for i in diskspace_all_value:
        sheet["A" + str(row_num)] = i[0]
        sheet["A" + str(row_num)].font = Font(size=9)
        sheet["B" + str(row_num)] = i[1]
        sheet["B" + str(row_num)].font = Font(size=9)
        sheet["C" + str(row_num)] = i[2]
        sheet["C" + str(row_num)].font = Font(size=9)
        sheet["D" + str(row_num)] = i[3]
        sheet["D" + str(row_num)].font = Font(size=9)
        if float(sheet["D" + str(row_num)].value) > 80:
            for cell_list in sheet["A" + str(row_num): "G" + str(row_num)]:
                for cell in cell_list:
                    cell.fill = cell_color
        sheet["E" + str(row_num)] = i[4]
        sheet["E" + str(row_num)].font = Font(size=9)
        sheet["F" + str(row_num)] = i[5]
        sheet["F" + str(row_num)].font = Font(size=9)
        sheet["G" + str(row_num)] = i[6]
        sheet["G" + str(row_num)].font = Font(size=9)
        row_num = row_num + 1
    workbook.save(export_file)
    print("diskspace report, success")


def diskgroup_query(query_list):
    print("diskgroup report, waiting ......")
    for i in query_list:
        diskgroup(i["host_ip"], i["project_info"])
    row_num = 2
    workbook = openpyxl.load_workbook(export_file)
    sheet = workbook.get_sheet_by_name("diskgroup")
    for i in diskgroup_all_value:
        sheet["A" + str(row_num)] = i[0]
        sheet["A" + str(row_num)].font = Font(size=9)
        sheet["B" + str(row_num)] = i[1]
        sheet["B" + str(row_num)].font = Font(size=9)
        sheet["C" + str(row_num)] = i[2]
        sheet["C" + str(row_num)].font = Font(size=9)
        sheet["D" + str(row_num)] = i[3]
        sheet["D" + str(row_num)].font = Font(size=9)
        if float(sheet["D" + str(row_num)].value) > 80:
            for cell_list in sheet["A" + str(row_num): "G" + str(row_num)]:
                for cell in cell_list:
                    cell.fill = cell_color
        sheet["E" + str(row_num)] = i[4]
        sheet["E" + str(row_num)].font = Font(size=9)
        sheet["F" + str(row_num)] = i[5]
        sheet["F" + str(row_num)].font = Font(size=9)
        sheet["G" + str(row_num)] = i[6]
        sheet["G" + str(row_num)].font = Font(size=9)
        row_num = row_num + 1
    workbook.save(export_file)
    print("diskgroup report, success")


# create_excel()
# diskspaceThread = threading.Thread(target=diskspace_query, args=(csv_to_dict("./diskspace.csv"),))
# tablespaceThread = threading.Thread(target=tablespace_query, args=(csv_to_dict("./tablespace.csv"),))
# diskgroupThread = threading.Thread(target=diskgroup_query, args=(csv_to_dict("./diskgroup.csv"),))
# diskspaceThread.start()
# tablespaceThread.start()
# diskgroupThread.start()
# diskspaceThread.join()
# tablespaceThread.join()
# diskgroupThread.join()

'''
with open("/tmp/report.json", "r") as f:
    a = json.loads(f.read())
for key, value in a.items():
    print(key)
    print(value)
'''

zabbix_url = "http://192.168.0.100/zabbix/api_jsonrpc.php"
headers = {
    "Content-Type": "application/json-rpc"
}


def get_token():
    if os.path.exists("/tmp/zabbix_token_test.txt"):
        with open("/tmp/zabbix_token_test.txt", "r") as f:
            zabbix_token = f.read()
    else:
        post_data = {
            "jsonrpc": "2.0",
            "method": "user.login",
            "params": {
                "user": "Admin",
                "password": "zabbix"
            },
            "id": 1
        }
        rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                              method='POST')
        response = request.urlopen(rqs)
        dict_access_token = json.loads(response.read())
        zabbix_token = dict_access_token["result"]
        with open("/tmp/zabbix_token_test.txt", "w") as f1:
            f1.write(zabbix_token)
    return zabbix_token


def get_hosts():
    post_data = {
        "jsonrpc": "2.0",
        "method": "host.get",
        "params": {
            "output": [
                "hostid",
                "host"
            ],
            "selectGroups": [
                "groupid",
                "name"
            ],
        },
        "id": 2,
        "auth": get_token()
    }
    rqs = request.Request(url=zabbix_url, data=json.dumps(post_data).encode('utf-8'), headers=headers,
                          method='POST')
    try_times = 1
    while try_times < 4:
        try:
            response = request.urlopen(rqs)
            return json.loads(response.read())["result"]
        except Exception as e:
            print(e)
            time.sleep(5)
            print(str(try_times) + " times reconnecting ...")
            try_times += 1


'''
def get_tablespace_info(groups):
    groups = [i for i in groups if i.find("db") > 0]
    print(groups)


groups_selected = ['oracle_production_os', 'oracle_production_db', 'os_production', 'os_test', 'oracle_notused_db',
                   'oracle_notused_os', 'oracle_test_db', 'oracle_test_os', 'Zabbix servers', ]
get_tablespace_info(groups_selected)
print(groups_selected)
'''

'''
partition_list = ["boot", "tmp", "opt", "usr", "var", "overlay", "docker", "/", "/app"]
partition_exclude = ["boot", "tmp", "opt", "usr", "var", "overlay", "docker"]
for j2 in partition_exclude:
    partition_list = [j3 for j3 in partition_list if j3.find(j2) < 0]
print(partition_list)
'''

'''
with open("/home/cloud/PycharmProjects/zabbix_report/nginx_new_dr.txt", "r") as f:
    for i in f:
        rq = urllib.request.Request(url=i)
        ssl._create_default_https_context = ssl._create_unverified_context
        try:
            response = urllib.request.urlopen(rq)
            print(i)
            with open("/home/cloud/PycharmProjects/zabbix_report/nginx_new_dr.txt", "a") as fw:
                fw.write(i)
        except Exception as e:
            print("url 404")
'''


def url_sort():
    url = []
    with open("/home/cloud/PycharmProjects/zabbix_report/nginx_new_dr.txt", "r") as f1:
        url_group_list = []
        for x in f1:
            rq = urllib.request.Request(x)
            ssl._create_default_https_context = ssl._create_unverified_context
            try:
                response = urllib.request.urlopen(rq)
            except Exception as e:
                print("url is 404")
            html = response.read()
            encode_type = chardet.detect(html)
            html = html.decode(encode_type['encoding'])
            reg = r'<td>[0-9]{1,}.[0-9]{1,}.[0-9]{1,}.[0-9]{1,}:[0-9]{1,}</td>'
            url_reg = re.compile(reg)
            url_list = re.findall(url_reg, html)
            url_ip_port = []
            for y in url_list:
                y = y.rstrip("</td>").lstrip("<td>")
                y = "http://" + y + "/"
                url_ip_port.append(y)
            x = x.lstrip("http://").rstrip("/status").lstrip("https://")
            url_group_list.append({x: url_ip_port})
        return url_group_list


def url_uniq():
    url_dict_1 = []
    url_dict_2 = []
    url_uniq_list = []
    z = 0
    for x in url_dict_1:
        url_uniq_list.append([])
    for y in url_dict_2:
        if url_dict_1[x] == url_dict_2[y]:
            url_uniq_list[z].append(y)
            url_dict_2.pop(y)
    z += 1
    print(url_uniq_list)


'''
def read_file_to_list():
    a = []
    with open("/home/cloud/PycharmProjects/zabbix_report/nginx_new_dr.txt", "r") as f2:
        for x in f2:
            a.append((x.rstrip("\n")))
    return a
'''


def dict_sort():
    dict1 = {"a": "1", "b": "2", "c": "3", "d": "3", "e": "1", "f": "2"}
    list_value = set(dict1.values())
    dict_value = {}
    for u in list_value:
        dict_value.update({u: []})
    for k, v in dict1.items():
        if v in dict_value:
            dict_value[v].append(k)
    print(dict_value)


def top_10(limit):
    """
    x is int
    """
    a = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 10, 2, 19, 2, 5, 7, 1, 19, 10]
    b = [0 for v in range(limit)]
    for x in a:
        if x <= b[limit - 1]:
            continue
        elif x >= b[0]:
            del b[limit - 1]
            b.insert(0, x)
        else:
            for y in range(limit - 1):
                if b[y] >= x >= b[y + 1]:
                    del b[limit - 1]
                    b.insert(y + 1, x)
                    break
    print(b)


def sort(nums):
    for x in range(len(nums) - 1):
        for y in range(len(nums) - x - 1):
            if nums[y] > nums[y + 1]:
                nums[y], nums[y + 1] = nums[y + 1], nums[y]
    print(nums)


def top_bubble(nums):
    for x in range(10):
        for y in range(len(nums) - x - 1):
            if nums[y] < nums[y + 1]:
                nums[y], nums[y + 1] = nums[y + 1], nums[y]
    print(nums)


def list_remove():
    a = ["a", "a", "a", "a", "a", "a", "a", "a", "a", "a"]
    for x in a:
        if x == "a":
            a.remove(x)
    print(a)


def multithread():
    nginx_list = []
    hosts_problem_nginx = []
    if not os.path.exists(BASE_DIR + "/nginx_sort.json"):
        with open(BASE_DIR + "/nginx_sort.json", "w") as nginx_sort:
            nginx_sort.write(json.dumps(query_data_api.nginx_sort()))
    with open(BASE_DIR + "/nginx_sort.json", "r") as nginx_sort:
        url_group = json.loads(nginx_sort.read())
        for k, v in url_group.items():
            nginx_list.append(list(v))

    threads = []
    work_queue = queue.Queue(len(nginx_list))
    queue_lock = threading.Lock()
    for x in nginx_list:
        work_queue.put(x[0])

    def get_host_down(nginx_url_queue, host_list):
        while not nginx_url_queue.empty():
            nginx_url = nginx_url_queue.get()
            for p in query_data_api.get_problems_backend_ip(nginx_url):
                if not p["ip"] in [i["host"] for i in host_list]:
                    host_list.append(
                        {"host": p["ip"], "fall_counts": p["fall_counts"], "nginx": nginx_url, "ignore": "0"})

    for x in range(11):
        t = threading.Thread(target=get_host_down, args=(work_queue, hosts_problem_nginx,))
        t.start()
        threads.append(t)
    for t in threads:
        t.join()
    print(len(hosts_problem_nginx))


multithread()
